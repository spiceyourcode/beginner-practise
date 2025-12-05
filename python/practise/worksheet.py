import openpyxl as xl 
from openpyxl.chart import BarChart, Reference 

wb = xl.load_workbook('C:/Users/Administrator/Desktop/programming/mosh.xlsx')
sheet= wb['Sheet1']


for rows in range(2 , sheet.max_row +1):
    cell=sheet.cell(rows,2)
    cell2=sheet.cell(rows,3)
    
    product_value = cell.value
    price = cell2.value
    corrected_price= product_value*price
    
    corrected_price_cell= sheet.cell(rows,4)
    corrected_price_cell.value=corrected_price
    
values=Reference(sheet, min_row=1, max_row=sheet.max_row,min_col= 4,max_col=4)
chart =BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'c8')    
    
try:
    wb.save('C:/Users/Administrator/Desktop/programming/python/mosh.xlsx')
except PermissionError:
    print("The file is open please close it first")
    

# print('The file is open close it to make changes to it')
    